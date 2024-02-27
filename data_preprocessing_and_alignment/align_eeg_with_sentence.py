import mne
import numpy as np
import openpyxl
import csv
import argparse
from utils import read_eeg_brainvision




def align_eeg_with_sentence(eeg_path, novel_xlsx_path, text_embedding_path, montage_name='GSN-HydroCel-128'):
    '''

    :param eeg_path: BrainVision files of the eeg
    :param novel_xlsx_path: path to the corresponding run of the novel
    :param text_embedding_path: path to the corresponding run of the text embeddings
    :return: cut_eeg_data, texts, text_embeddings in alignment. cut_eeg_data is a list containing all
             the cut eeg divided by the markers.
    '''

    eeg, events, event_id = read_eeg_brainvision(eeg_path, montage_name)

    text_embeddings = np.load(text_embedding_path)

    wb = openpyxl.load_workbook(novel_xlsx_path)
    wsheet = wb.active
    texts = []



    for i in range(2, wsheet.max_row + 1):
        texts.append((wsheet.cell(row=i, column=1)).value)

    start_chapter = int(texts[0])


    if start_chapter < 10:
        start_marker = 'CH0' + str(start_chapter)
    else:
        start_marker = 'CH' + str(start_chapter)


    start_marker_id = event_id[start_marker]
    events_start_chapter_index = np.where(events[:, 2] == start_marker_id)[0][0]

    eeg_data = eeg.get_data()



    ROWS_id = event_id['ROWS']
    ROWE_id = event_id['ROWE']


    rows_onset = []
    rowe_onset = []

    for event in events[events_start_chapter_index:]:
        if event[2] == ROWS_id:
            rows_onset.append(event[0])

    for event in events[events_start_chapter_index:]:
        if event[2] == ROWE_id:
            rowe_onset.append(event[0])




    cut_eeg_data = []

    for i in range(0, len(rows_onset)):

        start_time = rows_onset[i]
        end_time = rowe_onset[i]

        cut_eeg_data.append(eeg_data[:, start_time:end_time])


    return cut_eeg_data, texts, text_embeddings



parser = argparse.ArgumentParser(description='Parameters that can be changed')
parser.add_argument('--eeg_path', type=str, default=r'sub-07_ses-LittlePrince_task-reading_run-01_eeg.vhdr')
parser.add_argument('--novel_xlsx_path', type=str, default=r'segmented_Chinense_novel_run_1.xlsx')
parser.add_argument('--text_embedding_path', type=str, default=r'LittlePrince_text_embedding_run_1.npy')


args = parser.parse_args()



cut_eeg_data, texts, text_embeddings = align_eeg_with_sentence(eeg_path=args.eeg_path, novel_xlsx_path=args.novel_xlsx_path, text_embedding_path=args.text_embedding_path)