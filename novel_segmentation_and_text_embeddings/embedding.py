from transformers import BertTokenizer, AutoModelForMaskedLM
import openpyxl
import os
import numpy as np
import torch
import argparse

os.environ['CURL_CA_BUNDLE'] = ''

from transformers import AutoTokenizer, AutoModelForMaskedLM

#加载预训练字典和分词方法
tokenizer = BertTokenizer.from_pretrained(
    pretrained_model_name_or_path='bert-base-chinese',  # 可选，huggingface 中的预训练模型名称或路径，默认为 bert-base-chinese
)

model = AutoModelForMaskedLM.from_pretrained("bert-base-chinese")

parser = argparse.ArgumentParser(description='Parameters that can be changed')

parser.add_argument('--Chinese_novel_path', type=str, default=r'../data/segmented_novel',
                    help='Path to the folder that contains the .xlsx files of the texts')
parser.add_argument('--run_num', type=int, default=7,
                    help='Number of runs of your experiment')
parser.add_argument('--save_path', type=str, default=r'../data/embeddings',
                    help='Path to save the embedding files')

args = parser.parse_args()


for i in range(args.run_num):
    novel_path = args.Chinese_novel_path + '/segmented_Chinense_novel_run_' + str(i+1) + '.xlsx'

    wb = openpyxl.load_workbook(novel_path)
    wsheet = wb.active
    texts = []

    for j in range(2, wsheet.max_row + 1):
        texts.append((wsheet.cell(row=j, column=1)).value)

    print(texts)

    embeddings = []
    for k in range(len(texts)):
        token = tokenizer.encode(texts[k], return_tensors='pt')
        embedding = model(token).logits
        embedding = torch.mean(embedding, dim=1)
        embeddings.append(embedding.detach().numpy())


    embeddings = np.array(embeddings)

    embeddings = embeddings.reshape(embeddings.shape[0], embeddings.shape[2])

    print(embeddings)


    np.save(args.save_path + '/text_embedding_run_' + str(i+1) + '.npy', embeddings)



