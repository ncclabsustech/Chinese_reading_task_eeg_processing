import mne
import numpy as np
import openpyxl
import csv

def read_eeg_brainvision(eeg_path, montage_name='GSN-HydroCel-128'):
    eeg = mne.io.read_raw_brainvision(eeg_path)
    events = []
    event_id = {}

    events_path = eeg_path.replace('eeg.vhdr', 'events.tsv')

    with open(events_path) as events_file:
        csv_reader = csv.reader(events_file, delimiter='\t')  # 使用csv.reader读取csvfile中的文件
        header = next(csv_reader)        # 读取第一行每一列的标题
        for row in csv_reader:  # 将csv 文件中的数据保存到data中
            events.append([int(row[4]), 0, int(row[3])])  # 选择某一列加入到data数组中
            if row[2] not in event_id.keys():
                event_id[row[2]] = int(row[3])

    events = np.array(events)

    # event_desc = {value: key for key, value in event_id.items()}
    #
    # annotations = mne.annotations_from_events(events, sfreq=eeg.info['sfreq'], event_desc=event_desc)
    #
    # eeg.set_annotations(annotations)

    montage = mne.channels.make_standard_montage(montage_name)
    eeg.set_montage(montage)

    return eeg, events, event_id



def align_eeg_with_sentence(eeg_path, novel_path, start_index, end_index, montage_name='GSN-HydroCel-128'):
    '''

    :param eeg_path: BrainVision
    :param novel_path:
    :param containPreface:
    :return:
    '''

    eeg, events, event_id = read_eeg_brainvision(eeg_path, montage_name)



    wb = openpyxl.load_workbook(novel_path)
    wsheet = wb.active
    texts = []

    for i in range(2, wsheet.max_row + 1):
        texts.append((wsheet.cell(row=i, column=1)).value)


    text_start_index = texts.index(str(start_index))
    text_end_index = texts.index(str(end_index+1))
    texts = texts[text_start_index:text_end_index]


    print(texts)

    eeg_data = eeg.get_data()



    ROWS_id = event_id['ROWS']
    ROWE_id = event_id['ROWE']
    if start_index < 10:
        start_chapter_mark = 'CH0' + str(start_index)
    else:
        start_chapter_mark = 'CH' + str(start_index)

    start_chapter_id = event_id[start_chapter_mark]
    events_start_chapter_index = np.where(events[:, 2] == start_chapter_id)[0][0]

    onset = []

    for event in events[events_start_chapter_index:]:
        if event[2] == ROWS_id or event[2] == ROWE_id:
            onset.append(event[0])
    print(len(onset))

    cut_eeg_data = []

    for i in range(0, len(onset), 2):

        start_time = onset[i]
        end_time = onset[i+1]

        cut_eeg_data.append(eeg_data[:, start_time:end_time])

    print(cut_eeg_data[1].shape)
    print(len(cut_eeg_data))



align_eeg_with_sentence(eeg_path='sub-07_ses-LittlePrince_task-reading_run-01_eeg.vhdr', novel_path=r'../data/segmented_novel/segmented_Chinense_novel.xlsx', start_index=1, end_index=4)