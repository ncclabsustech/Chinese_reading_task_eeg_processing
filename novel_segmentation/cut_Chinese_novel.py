import argparse
import re
import openpyxl
import os

'''
Author: Jianyu Zhang, Xinyu Mou

This is used for novel segmentation and format transformation for the Chinese novel you want to play.
'''

def delete_specific_element(str, element):
    """删除字符串中特定的元素"""
    segments = re.split(element, str)
    segments = list(filter(lambda x:x != element, segments))
    result = ''.join(segments)

    return result

def contain_leading_quotation(sentence):
    if '“' in sentence:
        return True
    return False

def contain_back_quotation(sentence):
    if '”' in sentence:
        return True
    return False


def merge_short_sentences(segments):
    """将一个句子中被切分的过短的句子拼起来"""
    results = []
    results.append(segments[0])
    for i in range(1, len(segments)):
        if len(results[-1]) + len(segments[i]) <= 10:
            results[-1] += segments[i]
        else:
            results.append(segments[i])
    return results

def insert_element_to_str(str, element, index):
    """将指定元素插入到字符串的指定位置"""
    str_list = list(str)
    str_list.insert(index, element)
    result = ''.join(str_list)
    return result


def calculate_length_without_punctuation_and_indexes(sentence):
    """计算一个句子中除标点以外的长度和所有非标点位置的坐标"""
    punctuations = ['\n', '。', '，', '！', '？', '：', '；', '“', '”', '、', '《', '》', '.', '（', '）', '…', '·']
    sentence_list = list(sentence)
    length_without_punctuation = 0
    indexes = []
    for index, char in enumerate(sentence_list):
        if char not in punctuations:
            length_without_punctuation += 1
            indexes.append(index)

    return length_without_punctuation, indexes



def cut_paragraph(paragraph):
    """将文章切分完整的句子"""
    # 先切分整句
    sentences = re.split(r"(。|！|？|”|；)", paragraph)
    # 将单独的标点拼起来
    sentences = [''.join(i) for i in zip(sentences[0::2], sentences[1::2])]

    # 将标点移到正确的位置
    for i in range(len(sentences)):
        if sentences[i][0] in ['。', '！', '？', '”', '；']:
            sentences[i - 1] += sentences[i][0]
            sentences[i] = sentences[i][1:]

    # 去除掉空字符串
    sentences = list(filter(lambda x:x != '', sentences))
    sentences = [i.strip() for i in sentences]

    # 去除掉字符串中的\n
    sentences = [delete_specific_element(i, '\n') for i in sentences]

    # 去掉字符串中的空格
    sentences = [delete_specific_element(i, ' ') for i in sentences]

    # 将双引号不在同一个字符串的重新拼回来
    results = []
    isOneSentence = False
    for i in range(len(sentences)):
        # 前引号和后引号都有
        if contain_leading_quotation(sentences[i]) and contain_back_quotation(sentences[i]):
            results.append(sentences[i])
        # 只有前引号，代表后面有句子要添加进来
        elif contain_leading_quotation(sentences[i]) and not contain_back_quotation(sentences[i]):
            results.append(sentences[i])
            isOneSentence = True
        # 只有后引号，代表添加结束
        elif contain_back_quotation(sentences[i]):
            results[-1] += sentences[i]
            isOneSentence = False
        # 没有引号，且在引号中的句子
        elif isOneSentence == True:
            results[-1] += sentences[i]
        # 没有引号，且没有在引号中的句子
        else:
            results.append(sentences[i])

    return results


def cut_sentences(sentences):
    """检查每个句子是否超过十个字，如果超过了，就按逗号划分开"""
    results = []
    for i in range(len(sentences)):
        if len(sentences[i]) <= 10:
            results.append(sentences[i])
        else:
            segments = re.split(r"(，|：)", sentences[i])
            segments.append("")
            #print(segments)
            # 将单独的标点拼起来

            segments = [''.join(i) for i in zip(segments[0::2], segments[1::2])]

            #print(segments)
            # 将标点移到正确的位置
            for i in range(len(segments)):
                if segments[i][0] in ['，', '：']:
                    segments[i - 1] += segments[i][0]
                    segments[i] = segments[i][1:]
            # 去除掉空字符串
            segments = list(filter(lambda x: x != '', segments))
            segments = [k.strip() for k in segments]
            #print(segments)
            segments = merge_short_sentences(segments)
            for j in range(len(segments)):
                results.append(segments[j])

    return results


def arrange_sentences_within_30_words(sentences):
    """得到屏幕显示时每一帧的文字（每帧不超过30字）"""
    results = []
    results.append(sentences[0])

    # 将短句按屏幕所能容纳的量整合起来，假定每一帧不超过30个字，且每行最多10个字
    for i in range(1, len(sentences)):
        length_wiithout_punctuation_last, _ = calculate_length_without_punctuation_and_indexes(results[-1])
        length_wiithout_punctuation_new, _ = calculate_length_without_punctuation_and_indexes(sentences[i])


        if sentences[i] in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40']:
            results.append(sentences[i])
        elif sentences[i-1] in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40']:
            results.append(sentences[i])
        elif length_wiithout_punctuation_last + length_wiithout_punctuation_new < 31:
            results[-1] += sentences[i]
        else:
            results.append(sentences[i])

   # 将每一句话每十个字插入一个换行符
    for i in range(len(results)):
        sentence_length_without_punctuation, indexes_of_non_punctuation = calculate_length_without_punctuation_and_indexes(results[i])
        row_num = sentence_length_without_punctuation // 10
        for j in range(row_num):
            index = indexes_of_non_punctuation[(j+1)*10-1] + 1 + j
            results[i] = insert_element_to_str(results[i], '\n', index)

    return results





def split_chapter_title(sentences):
    """将每个章节的标题单独成句"""
    chapter_num = 0
    for i in range(len(sentences)):
        if sentences[i].find('Ch' + str(round(chapter_num))) != -1:
            index = sentences[i].find('Ch' + str(round(chapter_num)))
            segments = [sentences[i][:(index)], str(round(chapter_num)), sentences[i][(index+len(str(round(chapter_num)))+2):]]
            sentences[i] = segments[0]
            sentences.insert(i+1, segments[1])
            sentences.insert(i+2, segments[2])
            chapter_num += 1
    sentences = list(filter(lambda x:x != '', sentences))

    return sentences


def repeat_sentences(sentences):
    """将句子按其除掉标点后的长度重复，方便在psychopy中高亮时换帧"""
    results = []
    indexes = []
    punctuations = ['\n', '。', '，', '！', '？', '：', '；', '“', '”', '、', '《', '》', '.', '·']
    for i in range(len(sentences)):
        sentence = sentences[i]
        sentence_list = list(sentence)
        length_without_punctuation = 0
        for index, char in enumerate(sentence_list):
            if char not in punctuations:

                indexes.append(index)
                length_without_punctuation += 1

        for j in range(length_without_punctuation):
            results.append(sentence)

    #print(list(results[1]))

    return results, indexes


def split_row(sentences):
    """将文字按psychopy中显示时的每一行分起来"""
    results = []
    for i in range(len(sentences)):
        sentence_list = sentences[i].split('\n')
        sentence_list = list(filter(lambda x: x != '\n' and x != '', sentence_list))
        for j in range(len(sentence_list)):
            results.append(sentence_list[j])

    # 将句首有标点的转移到上一句的最后
    punctuations = ['。', '，', '！', '？', '：', '；', '”', '、', '》', '.', '）', '…', '·']
    for i in range(len(results)):
        if results[i][0] in punctuations:
            results[i-1] += results[i][0]
            results[i] = results[i][1:]

    results = list(filter(lambda x:x != '', results))
    #print(results)
    return results


def split_preface_main_content(sentences, divide_nums):
    """将前言部分单独分离开，并将正文部分按章节划分成指定数量的部分"""
    # def get_breakpoints(n, m):
    #     if m <= 1 or n <= 0:
    #         return []

    #     breakpoints = []
    #     interval = n // m + 1
    #     for i in range(1, m):
    #         breakpoint_value = i * interval
    #         breakpoints.append(breakpoint_value)

    #     return breakpoints


    if '1' in sentences:
        first_chapter_index = sentences.index('1')

    else:
        first_chapter_index = len(sentences)

    preface = sentences[:first_chapter_index]
    preface = preface[1:]   # 去掉最开始的标号0

    main_content = sentences[first_chapter_index:]


    max_chapter = 0
    while str(round(max_chapter+1)) in main_content:
        max_chapter += 1


    cut_chapter = divide_nums
    for i in range(len(divide_nums)):
        if cut_chapter[i] + 1 > max_chapter:
            cut_chapter.pop(i)
    cut_indexes_last = [main_content.index(str(round(i+1))) for i in cut_chapter]
    cut_indexes_last.append(len(main_content)+1)


    main_content_parts = []
    cut_index_first = 0
    for i in cut_indexes_last:
        main_content_parts.append(main_content[cut_index_first:i])
        cut_index_first = i

    return preface, main_content_parts




def arrange_sentences_in_psychopy_requirement(sentences):
    """需要高亮的行在中间，上下各有一行作为背景"""
    results = []
    indexes = []
    main_row = []
    row_num = []
    for i in range(len(sentences)):
        if i == 0 and sentences[i] not in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40']:
            length_without_puntuation, indexes_of_non_punc = calculate_length_without_punctuation_and_indexes(
                sentences[i])
            for k in range(length_without_puntuation):
                results.append(sentences[i] + '\n' + sentences[i + 1])
                indexes.append(indexes_of_non_punc[k])
                main_row.append(0)
                row_num.append(2)
        elif i == len(sentences) - 1:
            length_without_puntuation, indexes_of_non_punc = calculate_length_without_punctuation_and_indexes(
                sentences[i])
            for k in range(length_without_puntuation):
                results.append(sentences[i-1] + '\n' + sentences[i])
                indexes.append(indexes_of_non_punc[k])
                main_row.append(1)
                row_num.append(2)
        elif sentences[i] in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40']:
            results.append(sentences[i])
            indexes.append(0)
            main_row.append(0)
            row_num.append(1)
        elif sentences[i-1] in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40']:
            length_without_puntuation, indexes_of_non_punc = calculate_length_without_punctuation_and_indexes(sentences[i])
            for k in range(length_without_puntuation):
                results.append(sentences[i] + '\n' + sentences[i+1])
                indexes.append(indexes_of_non_punc[k])
                main_row.append(0)
                row_num.append(2)
        elif sentences[i+1] in ['0', '1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', '14', '15', '16', '17', '18', '19', '20', '21', '22', '23', '24', '25', '26', '27', '28', '29', '30', '31', '32', '33', '34', '35', '36', '37', '38', '39', '40']:
            length_without_puntuation, indexes_of_non_punc = calculate_length_without_punctuation_and_indexes(
                sentences[i])
            for k in range(length_without_puntuation):
                results.append(sentences[i-1] + '\n' + sentences[i])
                indexes.append(indexes_of_non_punc[k])
                main_row.append(1)
                row_num.append(2)
        else:
            length_without_puntuation, indexes_of_non_punc = calculate_length_without_punctuation_and_indexes(
                sentences[i])
            for k in range(length_without_puntuation):
                results.append(sentences[i-1] + '\n' + sentences[i] + '\n' + sentences[i+1])
                indexes.append(indexes_of_non_punc[k])
                main_row.append(1)
                row_num.append(3)
    return results, indexes, main_row, row_num


def save_to_xlsx(file_path, file_name, text, indexes=None, main_row=None, row_num=None):
    workbook = openpyxl.Workbook()
    sheet = workbook.active


    if not os.path.isdir(file_path):
        os.makedirs(file_path)

    if indexes is not None:
        for i, content in enumerate(zip(text, indexes, main_row, row_num)):
            sheet.cell(row=i + 2, column=1, value=content[0])
            sheet.cell(row=i + 2, column=2, value=content[1])
            sheet.cell(row=i + 2, column=3, value=content[2])
            sheet.cell(row=i + 2, column=4, value=content[3])

        sheet.cell(row=1, column=1, value='Chinese_text')
        sheet.cell(row=1, column=2, value='index')
        sheet.cell(row=1, column=3, value='main_row')
        sheet.cell(row=1, column=4, value='row_num')

        workbook.save(file_path + file_name)
    else:
        for i, content in enumerate(text):
            sheet.cell(row=i + 2, column=1, value=content)

        sheet.cell(row=1, column=1, value='Chinese_text')
        workbook.save(file_path + file_name)

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Parameters that can be changed in this experiment')
    
    parser.add_argument('--Chinese_novel_path', type=str, default=r'../data/novel/xiaowangzi_main_text.txt', help='Path to your .txt Chinese novel content')
    parser.add_argument('--divide_nums', type=str, default='4, 8, 12, 16, 20, 24', help='Breakpoints which you want to divide your novel (comma-separated)')
    args = parser.parse_args()

    divide_num_list = args.divide_nums.split(',')
  
    divide_num_list = [int(num) for num in divide_num_list]

    args.divide_nums = divide_num_list
    

    with open(args.Chinese_novel_path, encoding='utf-8') as file:
        text = file.read()
        result = cut_paragraph(text)
        result = cut_sentences(result)
        result = split_chapter_title(result)
        result = arrange_sentences_within_30_words(result)
        result = split_row(result)


        # 存psychopy需要用到的
        preface, main_content_parts = split_preface_main_content(result, args.divide_nums)

        preface_text, preface_indexes, preface_main_row, preface_row_num = arrange_sentences_in_psychopy_requirement(
            preface)

        save_to_xlsx(r'../data/segmented_novel', r'/segmented_Chinense_novel_preface_display.xlsx', preface_text,
                     preface_indexes, preface_main_row, preface_row_num)

        for i, main_content_part in enumerate(main_content_parts):
            text, indexes, main_row, row_num = arrange_sentences_in_psychopy_requirement(main_content_part)
            file_name = r'/segmented_Chinense_novel_run_' + str(round(i + 1)) + '_display.xlsx'
            save_to_xlsx(r'../data/segmented_novel', file_name, text, indexes, main_row, row_num)



        # 存用于检索的

        result_without_punc = []
        for row in result:
            length_without_punc, _ = calculate_length_without_punctuation_and_indexes(row)
            if length_without_punc != 0:
                result_without_punc.append(row)


        save_to_xlsx(r'../data/segmented_novel', r'/segmented_Chinense_novel.xlsx', result_without_punc[1:])

        preface_without_punc, main_content_parts_without_punc = split_preface_main_content(result_without_punc, args.divide_nums)



        save_to_xlsx(r'../data/segmented_novel', r'/segmented_Chinense_novel_preface.xlsx', preface_without_punc)

        for i, content_without_punc in enumerate(main_content_parts_without_punc):
            filename = r'/segmented_Chinense_novel_run_' + str(i+1) + '.xlsx'
            save_to_xlsx(r'../data/segmented_novel', filename, content_without_punc)






